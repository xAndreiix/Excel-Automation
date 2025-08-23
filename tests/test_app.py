import openpyxl
from app import process_workbook


def test_process_workbook(tmp_path):
    file = tmp_path / "test.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    ws.append(["Date", "Description", "Amount"])
    ws.append(["2025-01-05", "Item A", 100])
    ws.append(["2025-01-06", "Item B", 200])
    wb.save(file)

    process_workbook(file)

    wb2 = openpyxl.load_workbook(file)
    ws2 = wb2["Sheet1"]

    assert ws2.cell(2, 4).value == 90
    assert ws2.cell(3, 4).value == 180

    assert len(ws2._charts) > 0
